#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
腾讯云转码工具
用于将 Excel 文件中第四列的 Base64 编码数据解码，并新增第五列显示解码后的内容
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import base64
import os
import sys
from datetime import datetime
from typing import Optional, Tuple
import json
import subprocess
import platform

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    import xlwt
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def get_resource_path(relative_path):
    """
    获取资源文件的绝对路径
    支持开发环境和打包后的环境
    """
    try:
        # PyInstaller 打包后的临时文件夹
        base_path = sys._MEIPASS
    except Exception:
        # 开发环境
        base_path = os.path.abspath(os.path.dirname(__file__))
    
    return os.path.join(base_path, relative_path)


def open_file_with_system(file_path):
    """使用系统默认程序打开文件"""
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', file_path))
        elif platform.system() == 'Windows':
            os.startfile(file_path)
        else:  # Linux
            subprocess.call(('xdg-open', file_path))
        return True
    except Exception as e:
        return False


class TencentDecodeTool:
    def __init__(self, root):
        self.root = root
        self.root.title("腾讯云转码工具")
        self.root.geometry("900x850")
        
        # 选中的文件路径
        self.selected_file_path = None
        
        # 创建界面
        self.create_widgets()
    
    def create_widgets(self):
        """创建界面组件"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 工具说明区域
        info_frame = ttk.LabelFrame(main_frame, text="工具说明", padding="10")
        info_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        info_text = (
            "本工具用于处理腾讯云相关的 Excel 数据文件。\n"
            "功能：将 Excel 文件第四列（数据列）的 Base64 编码内容进行解码，"
            "并在第五列显示解码后的内容。\n"
            "支持格式：.xlsx 和 .xls\n"
            "输出文件：自动生成，格式为 '原文件名_转码_时间戳.xlsx'"
        )
        info_label = ttk.Label(info_frame, text=info_text, justify=tk.LEFT, foreground="blue")
        info_label.grid(row=0, column=0, sticky=tk.W)
        
        # 操作指引区域
        guide_frame = ttk.LabelFrame(main_frame, text="操作指引", padding="10")
        guide_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        guide_text = (
            "1. 点击 '选择 Excel 文件' 按钮，选择要处理的文件\n"
            "2. 确认文件路径显示正确\n"
            "3. 点击 '开始转码' 按钮执行处理\n"
            "4. 处理完成后，查看处理结果和输出文件路径\n"
            "注意：Excel 文件应包含表头（时间、通讯类型、Topic、数据），"
            "数据从第二行开始\n"
            "💡 提示：可以点击下方 '查看示例文件' 按钮查看转换前后的示例"
        )
        guide_label = ttk.Label(guide_frame, text=guide_text, justify=tk.LEFT, foreground="green")
        guide_label.grid(row=0, column=0, sticky=tk.W)
        
        # 文件选择区域
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # 文件路径显示
        self.file_path_label = ttk.Label(file_frame, text="未选择文件", foreground="gray")
        self.file_path_label.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 10))
        
        # 选择文件按钮
        self.select_file_btn = ttk.Button(
            file_frame, 
            text="选择 Excel 文件", 
            command=self.select_file
        )
        self.select_file_btn.grid(row=1, column=0, sticky=tk.W)
        
        # 处理按钮
        self.process_btn = ttk.Button(
            file_frame,
            text="开始转码",
            command=self.process_file,
            state='disabled'
        )
        self.process_btn.grid(row=1, column=1, sticky=tk.W, padx=(10, 0))
        
        # 示例文件区域
        example_frame = ttk.LabelFrame(main_frame, text="示例文件", padding="10")
        example_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        example_info = ttk.Label(
            example_frame, 
            text="查看转换前后的示例文件，帮助理解工具功能：",
            foreground="gray"
        )
        example_info.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        
        # 查看转换前示例按钮
        self.view_before_btn = ttk.Button(
            example_frame,
            text="查看转换前示例",
            command=self.view_example_before
        )
        self.view_before_btn.grid(row=1, column=0, sticky=tk.W, padx=(0, 5))
        
        # 查看转换后示例按钮
        self.view_after_btn = ttk.Button(
            example_frame,
            text="查看转换后示例",
            command=self.view_example_after
        )
        self.view_after_btn.grid(row=1, column=1, sticky=tk.W)
        
        # 结果显示区域
        result_frame = ttk.LabelFrame(main_frame, text="处理结果", padding="10")
        result_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 结果文本框（带滚动条）
        self.result_text = scrolledtext.ScrolledText(result_frame, height=15, width=80, wrap=tk.WORD)
        self.result_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置权重
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # 状态栏
        self.status_label = ttk.Label(main_frame, text="就绪", relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
    
    def select_file(self):
        """选择 Excel 文件"""
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx *.xls"),
                ("Excel 2007+", "*.xlsx"),
                ("Excel 2003", "*.xls"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_path:
            self.selected_file_path = file_path
            # 显示文件路径（如果太长则截断）
            display_path = file_path if len(file_path) <= 80 else "..." + file_path[-77:]
            self.file_path_label.config(text=f"已选择: {display_path}", foreground="green")
            self.process_btn.config(state='normal')
            self.update_status(f"已选择文件: {os.path.basename(file_path)}")
            self.log_result(f"✅ 已选择文件: {file_path}\n")
        else:
            self.selected_file_path = None
            self.file_path_label.config(text="未选择文件", foreground="gray")
            self.process_btn.config(state='disabled')
            self.update_status("未选择文件")
    
    def update_status(self, message: str):
        """更新状态栏"""
        self.status_label.config(text=f"状态: {message}")
        self.root.update_idletasks()
    
    def log_result(self, message: str):
        """记录处理结果"""
        self.result_text.insert(tk.END, message)
        self.result_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_result(self):
        """清空结果"""
        self.result_text.delete(1.0, tk.END)
    
    def decode_base64(self, encoded_str: str) -> Tuple[bool, str]:
        """
        解码 Base64 字符串
        
        Returns:
            (success, decoded_str): 成功标志和解码后的字符串
        """
        try:
            # 去除可能的空白字符
            encoded_str = encoded_str.strip()
            if not encoded_str:
                return False, "空字符串"
            
            # Base64 解码
            decoded_bytes = base64.b64decode(encoded_str)
            decoded_str = decoded_bytes.decode('utf-8')
            
            # 尝试格式化为 JSON（如果是 JSON 格式）
            try:
                json_obj = json.loads(decoded_str)
                decoded_str = json.dumps(json_obj, ensure_ascii=False, indent=2)
            except (json.JSONDecodeError, ValueError):
                # 不是 JSON 格式，直接返回原始字符串
                pass
            
            return True, decoded_str
        except Exception as e:
            return False, f"解码失败: {str(e)}"
    
    def process_xlsx(self, input_path: str, output_path: str) -> Tuple[bool, str, int]:
        """
        处理 .xlsx 文件
        
        Returns:
            (success, message, processed_count): 成功标志、消息、处理的行数
        """
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl 库未安装，无法处理 .xlsx 文件", 0
        
        try:
            # 读取工作簿
            workbook = openpyxl.load_workbook(input_path)
            sheet = workbook.active
            
            processed_count = 0
            error_count = 0
            
            # 检查是否有表头
            if sheet.max_row < 1:
                return False, "Excel 文件为空", 0
            
            # 检查是否需要添加第五列表头
            if sheet.max_column < 4:
                return False, "Excel 文件列数不足，至少需要4列（时间、通讯类型、Topic、数据）", 0
            
            # 如果只有4列，添加第五列表头
            if sheet.max_column == 4:
                header_cell = sheet.cell(row=1, column=5)
                header_cell.value = "解码数据"
            
            # 从第二行开始处理（第一行是表头）
            for row_idx in range(2, sheet.max_row + 1):
                # 获取第四列的数据（索引为4，即D列）
                data_cell = sheet.cell(row=row_idx, column=4)
                base64_str = str(data_cell.value) if data_cell.value else ""
                
                if not base64_str or base64_str.strip() == "":
                    # 空数据，跳过
                    continue
                
                # 解码 Base64
                success, decoded_str = self.decode_base64(base64_str)
                
                if success:
                    # 写入第五列（索引为5，即E列）
                    decoded_cell = sheet.cell(row=row_idx, column=5)
                    decoded_cell.value = decoded_str
                    processed_count += 1
                else:
                    error_count += 1
                    self.log_result(f"⚠️  第 {row_idx} 行解码失败: {decoded_str}\n")
            
            # 保存文件
            workbook.save(output_path)
            workbook.close()
            
            message = f"处理完成！成功: {processed_count} 行，失败: {error_count} 行"
            return True, message, processed_count
            
        except Exception as e:
            return False, f"处理 .xlsx 文件时出错: {str(e)}", 0
    
    def process_xls(self, input_path: str, output_path: str) -> Tuple[bool, str, int]:
        """
        处理 .xls 文件
        
        Returns:
            (success, message, processed_count): 成功标志、消息、处理的行数
        """
        if not XLRD_AVAILABLE:
            return False, "xlrd/xlwt 库未安装，无法处理 .xls 文件", 0
        
        try:
            # 读取工作簿
            workbook = xlrd.open_workbook(input_path)
            sheet = workbook.sheet_by_index(0)
            
            # 创建新的工作簿用于写入
            output_workbook = xlwt.Workbook()
            output_sheet = output_workbook.add_sheet(sheet.name)
            
            processed_count = 0
            error_count = 0
            
            # 检查是否有数据
            if sheet.nrows < 1:
                return False, "Excel 文件为空", 0
            
            # 检查列数
            if sheet.ncols < 4:
                return False, "Excel 文件列数不足，至少需要4列（时间、通讯类型、Topic、数据）", 0
            
            # 复制所有数据并添加第五列
            for row_idx in range(sheet.nrows):
                # 复制所有原有列
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    output_sheet.write(row_idx, col_idx, cell_value)
                
                # 如果是第一行（表头），添加第五列表头（如果还没有）
                if row_idx == 0:
                    if sheet.ncols == 4:
                        output_sheet.write(row_idx, 4, "解码数据")
                else:
                    # 从第二行开始处理数据
                    base64_str = str(sheet.cell_value(row_idx, 3)) if sheet.ncols > 3 else ""
                    
                    if base64_str and base64_str.strip() != "":
                        # 解码 Base64
                        success, decoded_str = self.decode_base64(base64_str)
                        
                        if success:
                            output_sheet.write(row_idx, 4, decoded_str)
                            processed_count += 1
                        else:
                            error_count += 1
                            self.log_result(f"⚠️  第 {row_idx + 1} 行解码失败: {decoded_str}\n")
            
            # 保存文件
            output_workbook.save(output_path)
            
            message = f"处理完成！成功: {processed_count} 行，失败: {error_count} 行"
            return True, message, processed_count
            
        except Exception as e:
            return False, f"处理 .xls 文件时出错: {str(e)}", 0
    
    def generate_output_filename(self, input_path: str) -> str:
        """生成输出文件名"""
        # 获取文件目录和文件名（不含扩展名）
        file_dir = os.path.dirname(input_path)
        file_name = os.path.splitext(os.path.basename(input_path))[0]
        
        # 生成时间戳
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 生成输出文件名
        output_filename = f"{file_name}_转码_{timestamp}.xlsx"
        output_path = os.path.join(file_dir, output_filename)
        
        return output_path
    
    def view_example_before(self):
        """查看转换前的示例文件"""
        example_path = get_resource_path("resources/上下行命令.xlsx")
        
        if not os.path.exists(example_path):
            messagebox.showerror(
                "错误", 
                f"示例文件不存在：\n{example_path}\n\n"
                "请确保资源文件已正确打包到应用中。"
            )
            return
        
        if open_file_with_system(example_path):
            self.log_result(f"📄 已打开转换前示例文件: {example_path}\n")
            self.update_status("已打开转换前示例文件")
        else:
            messagebox.showerror("错误", f"无法打开文件：\n{example_path}")
    
    def view_example_after(self):
        """查看转换后的示例文件"""
        # 尝试找到最新的转换后示例文件
        resource_dir = os.path.dirname(get_resource_path("resources/上下行命令.xlsx"))
        example_files = [
            get_resource_path("resources/上下行命令_转码_20251212_100955.xlsx"),
        ]
        
        # 查找存在的示例文件
        example_path = None
        for path in example_files:
            if os.path.exists(path):
                example_path = path
                break
        
        if not example_path:
            messagebox.showerror(
                "错误", 
                f"示例文件不存在\n\n"
                "请确保资源文件已正确打包到应用中。"
            )
            return
        
        if open_file_with_system(example_path):
            self.log_result(f"📄 已打开转换后示例文件: {example_path}\n")
            self.update_status("已打开转换后示例文件")
        else:
            messagebox.showerror("错误", f"无法打开文件：\n{example_path}")
    
    def process_file(self):
        """处理选中的文件"""
        if not self.selected_file_path:
            messagebox.showwarning("警告", "请先选择要处理的 Excel 文件")
            return
        
        if not os.path.exists(self.selected_file_path):
            messagebox.showerror("错误", "选择的文件不存在")
            return
        
        # 检查文件扩展名
        file_ext = os.path.splitext(self.selected_file_path)[1].lower()
        if file_ext not in ['.xlsx', '.xls']:
            messagebox.showerror("错误", f"不支持的文件格式: {file_ext}\n请选择 .xlsx 或 .xls 文件")
            return
        
        # 检查必要的库
        if file_ext == '.xlsx' and not OPENPYXL_AVAILABLE:
            messagebox.showerror("错误", "无法处理 .xlsx 文件：openpyxl 库未安装\n请运行: pip install openpyxl")
            return
        
        if file_ext == '.xls' and not XLRD_AVAILABLE:
            messagebox.showerror("错误", "无法处理 .xls 文件：xlrd/xlwt 库未安装\n请运行: pip install xlrd xlwt")
            return
        
        # 清空之前的结果
        self.clear_result()
        
        # 生成输出文件名
        output_path = self.generate_output_filename(self.selected_file_path)
        
        # 禁用按钮
        self.select_file_btn.config(state='disabled')
        self.process_btn.config(state='disabled')
        
        self.update_status("正在处理...")
        self.log_result(f"📂 输入文件: {self.selected_file_path}\n")
        self.log_result(f"📂 输出文件: {output_path}\n")
        self.log_result(f"📋 开始处理...\n\n")
        
        try:
            # 根据文件类型选择处理方法
            if file_ext == '.xlsx':
                success, message, count = self.process_xlsx(self.selected_file_path, output_path)
            else:  # .xls
                success, message, count = self.process_xls(self.selected_file_path, output_path)
            
            if success:
                self.log_result(f"✅ {message}\n\n")
                self.log_result(f"📁 输出文件已保存到: {output_path}\n")
                self.update_status(f"处理完成 - 成功处理 {count} 行")
                messagebox.showinfo("处理完成", f"{message}\n\n输出文件:\n{output_path}")
            else:
                self.log_result(f"❌ {message}\n")
                self.update_status("处理失败")
                messagebox.showerror("处理失败", message)
        
        except Exception as e:
            error_msg = f"处理过程中发生异常: {str(e)}"
            self.log_result(f"❌ {error_msg}\n")
            self.update_status("处理异常")
            messagebox.showerror("异常", error_msg)
        
        finally:
            # 重新启用按钮
            self.select_file_btn.config(state='normal')
            self.process_btn.config(state='normal')


def main():
    """主函数"""
    # 检查必要的库
    missing_libs = []
    if not OPENPYXL_AVAILABLE:
        missing_libs.append("openpyxl (用于处理 .xlsx 文件)")
    if not XLRD_AVAILABLE:
        missing_libs.append("xlrd 和 xlwt (用于处理 .xls 文件)")
    
    if missing_libs:
        print("⚠️  警告: 以下库未安装，某些功能可能无法使用:")
        for lib in missing_libs:
            print(f"  - {lib}")
        print("\n请运行以下命令安装:")
        if not OPENPYXL_AVAILABLE:
            print("  pip install openpyxl")
        if not XLRD_AVAILABLE:
            print("  pip install xlrd xlwt")
        print()
    
    root = tk.Tk()
    app = TencentDecodeTool(root)
    root.mainloop()


if __name__ == '__main__':
    main()
